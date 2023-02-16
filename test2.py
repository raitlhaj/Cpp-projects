from random import *
from pathlib import *
import openpyxl as xl
from openpyxl.chart import BarChart,Reference


w=xl.load_workbook("Classeur1.xlsx")
sh=w['Feuil1']
c=sh.cell(1,1)

members=['Rachid','Ali','Dija','Ahmed']

#for i in range(3):
#    print(randint(10,20))

#print(choice(members))

path=Path()

for file in path.glob('*.py'):
    print(file)

for row in range(1,sh.max_row+1):
    val=sh.cell(row,3).value
    correctVa=val*5
    correctCell=sh.cell(row,4)
    correctCell.value=correctVa

values=Reference(sh,min_row=2,max_row=sh.max_row,min_col=4,max_col=4)
chart=BarChart()
chart.add_data(values)
sh.add_chart(chart,'e3')

w.save('Classeur1.xlsx')